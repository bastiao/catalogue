# Copyright (C) 2014 Ricardo Ribeiro and Universidade de Aveiro
#
# Authors: Ricardo Ribeiro <ribeiro.r@ua.pt>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#

from django.core.management.base import BaseCommand, CommandError

from django.contrib.auth.models import User

from questionnaire.models import Questionnaire, Choice, Question, QuestionSet

import os
import re

from openpyxl import Workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell

class Command(BaseCommand):

    args = '<database_slug> <file_path>'
    help = 'Export the questionnaire to excel'

    # dependency questions need to be able to translate question numbers into excel line numbers.
    __number_map = {}

    def __boolean_to_string(self, value):
        if value == True:
            return 'yes'
        elif value == False:
            return 'no'

        return 'error'


    def __setDefaultStyle(self, _cell):
        _cell.style.font.name = 'Verdana'
        _cell.style.font.size = 8
        _cell.style.alignment.wrap_text = True

    def __setBold(self, _cell):
        _cell.style.font.bold = True

    def __setHeader(self, _cell):
        _cell.style.font.bold = True

        # Cell background color
        _cell.style.fill.fill_type = Fill.FILL_SOLID
        _cell.style.fill.start_color.index = 'FFCCCCCC'
        _cell.style.alignment.horizontal = 'center'

    def __setColumnSizes(self, ws, sizes):
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

        for i in xrange(len(columns)):
            ws.column_dimensions[columns[i]].width = sizes[i]
    def __getChoices(self, question):
        retstring = ""

        choices = Choice.objects.filter(question=question)

        for choice in choices:
            retstring+=choice.value+'|'

        if(len(retstring) > 1):
            return retstring[:-1]

        return retstring

    def __getChoiceNumber(self, parent, option):
        yesno_questions = ['choice-yesno','choice-yesnocomment','choice-yesnodontknow']

        if parent.type in yesno_questions:
            if option.lower() == 'yes':
                return "1"
            elif option.lower() == 'no':
                return "2"
            elif option.lower() == 'dontknow':
                return "3"

            return "error"
        else:
            return Choice.objects.get(question=parent, value=option).sortid


    def __processDependencies(self, question):

        try:
            if (question.checks == None):
                question.checks = ""
            valid = re.search('.*dependent="([0-9\.]+),(.*)".*', question.checks, re.IGNORECASE)

            if valid:
                number = valid.group(1)
                option = valid.group(2)

                (line, parent) = self.__number_map[number]

                optionid = self.__getChoiceNumber(parent, option)

                restring = '=ROW($'+str(line)+':$'+str(line)+') & "|'+str(optionid)+'"'

                return restring

            # No dependencies
            else:
                return ""
        except KeyError:
            print "-- ERROR: Couldn't find a mapping for question "+str(question.number)
            return "error"





    def __addQuestion(self, line, ws, question):
        self.__number_map[question.number] = (line, question)

        valid = re.search('(h[0-9])+\. (.*)', question.text_en, re.IGNORECASE)
        choice_types = ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform', 'choice-multiple-freeform-options']

        if valid:
            level = str(valid.group(1))
            text = str(valid.group(2))
            type = None

            choices = ''

            if question.category:
                type = 'Category'
            else:
                type = 'Question'

            if question.type in choice_types:
                choices = self.__getChoices(question)

            ws.append([
                    type,
                    text,
                    level,
                    question.type,
                    choices,
                    question.help_text,
                    question.tooltip,
                    question.slug_fk.slug1,
                    self.__processDependencies(question),
                    '',
                    '',
                ])

            if question.category:
                self.__setBold(ws.cell('B'+str(line)))

            return True

        return False

    def handle(self, *args, **options):
        self.__number_map = {}
        if len(args) == 2:
            slug = args[0]
            file_path = args[1]

            questionnaire = Questionnaire.objects.get(slug=slug)

            wb = Workbook()
            ws = wb.get_active_sheet()
            ws.title = "Questionnaire"

            ws.cell('A1').value = "Name"
            self.__setHeader(ws.cell('A1'))
            ws.cell('B1').value = slug
            self.__setBold(ws.cell('B1'))

            ws.append(['Type', 'Text/Question', 'Level/Number', 'Data Type', 'Value list', 'Help text/Description',
                       'Tooltip', 'Slug', 'Dependencies', 'Stats', 'Comments State'])

            for row in ws.range('A2:K2'):
                for cell in row:
                    self.__setHeader(cell)

            # for sanity, im keeping a pointer to the row im in...
            pointer = 3

            for questionset in questionnaire.questionsets():
                ws.append(['QuestionSet', questionset.text_en.replace('h1. ',''),
                            questionset.sortid, '', '', questionset.help_text.replace('<br />', '\n'),
                            self.__boolean_to_string(questionset.tooltip), '',
                            '', '', '' ])

                self.__setBold(ws.cell('A'+str(pointer)))
                self.__setBold(ws.cell('B'+str(pointer)))
                self.__setBold(ws.cell('C'+str(pointer)))

                pointer += 1

                for question in questionset.questions():
                    inserted = self.__addQuestion(pointer, ws, question)

                    if inserted:
                        pointer += 1
                    else:
                        print "-- ERROR PROCESSING QUESTION header for: "+str(question.text_en)
                        break

            self.__setColumnSizes(ws, [10, 30, 15, 15, 30, 30, 12, 15, 15, 7, 15])

            for row in ws.rows:
                for cell in row:
                    self.__setDefaultStyle(cell)

            wb.save(file_path)


        else:
            self.stdout.write('-- USAGE: \n    '+
                'python manage.py export_questionnaire <database_slug> <path_file>'+
                '\n\n')
